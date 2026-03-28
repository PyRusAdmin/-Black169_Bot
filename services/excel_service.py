# -*- coding: utf-8 -*-
import io
from openpyxl import Workbook
from loguru import logger


def write_users_to_excel(data: list) -> io.BytesIO:
    """
    Запись данных пользователей в Excel-файл в памяти
    
    :param data: Список словарей с данными пользователей
    :return: Буфер с Excel-файлом в формате BytesIO
    """
    wb = Workbook()  # создаем новый Excel-файл
    ws = wb.active  # активный лист в Excel-файле

    # Заголовок
    ws.append(["ID Telegram", "Имя", "Фамилия", "Username", "Дата регистрации"])

    # Данные
    for person in data:
        ws.append([
            person.get("id_telegram", ""),  # ID Telegram
            person.get("first_name", ""),  # Имя
            person.get("last_name", ""),  # Фамилия
            person.get("username", ""),  # Username
            person.get("updated_at", "")  # Дата регистрации
        ])

    """ 
    io.BytesIO — класс из стандартного модуля io в Python, который создаёт бинарный поток в памяти. Он ведёт себя 
    аналогично файлу, открытому в бинарном режиме, но данные хранятся в оперативной памяти (RAM) вместо на диске. 
    """

    buffer = io.BytesIO()  # буфер для Excel-файла
    wb.save(buffer)  # сохраняем Excel-файл в буфер
    buffer.seek(0)  # сдвигаем указатель в начало буфера
    logger.success("Excel-файл с пользователями сформирован в памяти")
    return buffer  # возвращаем буфер с Excel-файлом


def write_winners_to_excel(data: list) -> io.BytesIO:
    """
    Запись данных победителей «Колеса подарков» в Excel-файл в памяти
    
    :param data: Список словарей с данными победителей
    :return: Буфер с Excel-файлом в формате BytesIO
    """
    wb = Workbook()  # создаем новый Excel-файл
    ws = wb.active  # активный лист в Excel-файле

    # Заголовок
    ws.append(["ID Telegram", "ID QuickResto", "Приз", "Дата выигрыша"])

    # Данные
    for winner in data:
        ws.append([
            winner.get("id_telegram", ""),  # ID Telegram
            winner.get("id_quickresto", ""),  # ID QuickResto
            winner.get("bonus_name", ""),  # Приз
            winner.get("spun_at", "")  # Дата выигрыша
        ])

    """ 
    io.BytesIO — класс из стандартного модуля io в Python, который создаёт бинарный поток в памяти. Он ведёт себя 
    аналогично файлу, открытому в бинарном режиме, но данные хранятся в оперативной памяти (RAM) вместо на диске. 
    """

    buffer = io.BytesIO()  # буфер для Excel-файла
    wb.save(buffer)  # сохраняем Excel-файл в буфер
    buffer.seek(0)  # сдвигаем указатель в начало буфера
    logger.success("Excel-файл с победителями сформирован в памяти")
    return buffer  # возвращаем буфер с Excel-файлом


def write_registered_users_to_excel(data: list) -> io.BytesIO:
    """
    Запись данных зарегистрированных пользователей (кто отправил номер телефона) в Excel-файл в памяти
    
    :param data: Список словарей с данными зарегистрированных пользователей
    :return: Буфер с Excel-файлом в формате BytesIO
    """
    wb = Workbook()  # создаем новый Excel-файл
    ws = wb.active  # активный лист в Excel-файле

    # Заголовок
    ws.append([
        "ID Telegram", "ID QuickResto", "Телефон", "Фамилия", "Имя", "Отчество",
        "Дата рождения", "Бонусы", "Дата последнего визита", "Дата обновления"
    ])

    # Данные
    for user in data:
        ws.append([
            user.get("id_telegram", ""),  # ID Telegram
            user.get("id_quickresto", ""),  # ID QuickResto
            user.get("phone_telegram", ""),  # Телефон
            user.get("last_name", ""),  # Фамилия
            user.get("first_name", ""),  # Имя
            user.get("patronymic_name", ""),  # Отчество
            user.get("birthday_user", ""),  # Дата рождения
            user.get("user_bonus", ""),  # Бонусы
            user.get("date_of_visit", ""),  # Дата последнего визита
            user.get("updated_at", "")  # Дата обновления
        ])

    """ 
    io.BytesIO — класс из стандартного модуля io в Python, который создаёт бинарный поток в памяти. Он ведёт себя 
    аналогично файлу, открытому в бинарном режиме, но данные хранятся в оперативной памяти (RAM) вместо на диске. 
    """

    buffer = io.BytesIO()  # буфер для Excel-файла
    wb.save(buffer)  # сохраняем Excel-файл в буфер
    buffer.seek(0)  # сдвигаем указатель в начало буфера
    logger.success("Excel-файл с зарегистрированными пользователями сформирован в памяти")
    return buffer  # возвращаем буфер с Excel-файлом
